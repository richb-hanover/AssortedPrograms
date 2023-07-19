
# Thanks, ChatGPT!
# https://chat.openai.com/c/61bf6d71-67b2-4a1f-b5b0-7ecb3292b699
# Getting close 

import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_NUMBER_COMMA_SEPARATED2
from openpyxl.styles import Font, Alignment
from openpyxl.styles import numbers

# Set the folder path where the Excel files are located
folder_path = "/Users/richb/Documents/Businesses/PinnacleProject/Cabin Condo Finance-2023/QB Items for Claremont"

# Get the list of Excel files in the folder
excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# Iterate over each Excel file
for file_name in excel_files:
    # Remove the extension from the file name
    name_without_extension = os.path.splitext(file_name)[0]
    if (name_without_extension.find("modified") != -1) or (name_without_extension.find("~")!= -1):
        print("Skipping", name_without_extension)
        continue
    
    print("Processing", name_without_extension)
    
    # Create the file path
    file_path = os.path.join(folder_path, file_name)

    # Load the workbook
    workbook = load_workbook(filename=file_path)

    # Get the first sheet
    first_sheet = workbook.worksheets[0]

    try:
        # Delete the first sheet
        workbook.remove(first_sheet)
    except:
        print("No first sheet")
    
    # Get the (new) first sheet
    first_sheet = workbook.worksheets[0]

    # Set cell A1 to the file name without extension
    first_sheet['A1'] = name_without_extension
    # Set cell A1 to be left-adjusted
    first_sheet['A1'].alignment = Alignment(horizontal='left')
   
    # Change the font to Calibri 12 for all sheets
    font = Font(name='Calibri', size=12)
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            for cell in column:
                cell.font = font

    # Convert all numbers to currency format
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            # sheet.row_dimensions[row].height = 100
            for cell in row:
                # if isinstance(cell.value, (int, float)):
                    # Set the number format to currency with no decimal places
                cell.number_format = '$#,##0'
    
                    # Convert negative numbers to parentheses format
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        cell.number_format = '($#,##0)'
                        
    # Save the modified workbook
    modified_file_path = os.path.join(folder_path, f"modified_{name_without_extension}.xlsx")
    workbook.save(filename=modified_file_path)

    print(f"Processed file: {file_name}")
