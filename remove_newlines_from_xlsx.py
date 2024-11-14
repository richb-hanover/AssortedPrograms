# From ChatGPT
# do the same thing, except process all tabs on an Excel .xlsx file
#
# Usage: python remove_newlines_from_xlsx.py < input.xlsx > output.xlsx
#
# https://chat.openai.com/c/536fbd55-0a31-4429-955a-78dcb021d701

import openpyxl
import sys

!!! THIS DOES NOT WORK - mumbles something about not being a ZIP file

# Create a Workbook object from stdin
try:
    input_workbook = openpyxl.load_workbook(sys.stdin, read_only=True, data_only=True)
except Exception as e:
    sys.stderr.write(f"Error reading Excel file: {str(e)}")
    sys.exit(1)

# Create a new Workbook to store the cleaned data
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Iterate through all sheets in the input workbook
for sheet_name in input_workbook.sheetnames:
    input_sheet = input_workbook[sheet_name]
    output_sheet = output_workbook.create_sheet(title=sheet_name)

    # Iterate through rows and columns
    for row in input_sheet.iter_rows(values_only=True):
        cleaned_row = []
        for cell in row:
            if isinstance(cell, str):
                cleaned_cell = cell.replace('\n', ' ').replace('\r', ' ')
            else:
                cleaned_cell = cell
            cleaned_row.append(cleaned_cell)
        output_sheet.append(cleaned_row)

# Save the cleaned data to stdout
output_workbook.save(sys.stdout)
